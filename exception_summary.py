import pandas as pd
import os
from datetime import date
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ================= 配置区域 =================

base_path = r"C:\Users\zhangyaolong\Desktop\异常概括"
archive_sheet_name = "问题归档"  # 指定归档页的名字

# --- 板块 A: 自营采购退货 ---
files_self_run = [
    (r"C:\Users\zhangyaolong\Desktop\异常概括\每日\自营退货订单项目维度退货单.xlsx", "自营退货订单-项目维度退货单"),
    (r"C:\Users\zhangyaolong\Desktop\异常概括\每日\项目维度退货单逆向销售单.xlsx", "项目维度退货单-逆向销售单")
]

# --- 板块 B: 备件库退货 ---
files_spare_parts = [
    (r"C:\Users\zhangyaolong\Desktop\异常概括\每日\备件库退货订单项目维度退货单.xlsx", "备件库退货订单-项目维度退货单"),
    (r"C:\Users\zhangyaolong\Desktop\异常概括\每日\备件库项目维度退货单逆向销售单.xlsx", "项目维度退货单-逆向销售单")
]

# --- 板块 C: 平台零售订单 ---
files_platform_retail = [
    (r"C:\Users\zhangyaolong\Desktop\异常概括\每日\销售订单应收单.xlsx", "销售订单-应收单"),
    (r"C:\Users\zhangyaolong\Desktop\异常概括\每日\发货通知单销售出库单.xlsx", "发货通知单-销售出库单"),
    (r"C:\Users\zhangyaolong\Desktop\异常概括\每日\销售出库单应收单.xlsx", "销售出库单-应收单")
]


# ===========================================

def get_current_date_col():
    return date.today().strftime("%m/%d")
    # return "12/25"


def read_daily_data(file_config_list, current_date):
    """ 读取当天的原始文件 """
    df_list = []
    node_order = [item[1] for item in file_config_list if item[1]]

    print(f"--- 正在读取新数据 ---")
    for file_path, node_name in file_config_list:
        if not file_path or not os.path.exists(file_path):
            print(f"⚠️ 跳过 (未找到文件): {os.path.basename(file_path)}")
            continue
        try:
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                df = pd.read_excel(file_path)
            else:
                try:
                    df = pd.read_csv(file_path, encoding='gbk')
                except:
                    df = pd.read_csv(file_path, encoding='utf-8-sig')

            df.columns = [c.strip() for c in df.columns]
            df['环节'] = node_name

            # 读取源表自带的 '异常备注'
            target_cols = ['异常类型', '异常描述', '异常信息', '异常备注']
            for col in target_cols:
                if col not in df.columns: df[col] = ''
                df[col] = df[col].fillna('').astype(str)

            df_list.append(df)
        except Exception as e:
            print(f"读取出错 {file_path}: {e}")

    if not df_list:
        return pd.DataFrame(), node_order

    df_all = pd.concat(df_list)

    # 分组时加入 '异常备注' (即源表备注)
    group_keys = ['环节', '异常类型', '异常描述', '异常信息', '异常备注']
    summary = df_all.groupby(group_keys).size().reset_index(name=current_date)
    summary = summary.rename(columns={'异常备注': '源表异常备注'})

    return summary, node_order


def load_all_sheets(master_file_path):
    """ 读取历史数据 """
    if not os.path.exists(master_file_path):
        return {}
    try:
        print(f"正在读取汇总表所有 Sheet...")
        all_sheets = pd.read_excel(master_file_path, sheet_name=None, engine='openpyxl')
        for name, df in all_sheets.items():
            if not df.empty:
                check_cols = ['异常类型', '异常描述', '异常信息', '源表异常备注',
                              '优先级', '计划', '责任人', '异常修复', '备注']
                for col in check_cols:
                    if col in df.columns:
                        df[col] = df[col].fillna('').astype(str)
        return all_sheets
    except Exception as e:
        print(f"⚠️ 历史文件读取异常: {e}")
        return {}


def adjust_excel_style(writer, sheet_name, df):
    """ 美化函数 """
    worksheet = writer.sheets[sheet_name]
    worksheet.row_dimensions[1].height = 25
    for row_idx in range(2, len(df) + 2):
        worksheet.row_dimensions[row_idx].height = 28
    for idx, col in enumerate(df.columns):
        series = df[col]
        max_len = len(str(col).encode('gbk'))
        for item in series.head(100):
            try:
                length = len(str(item).encode('gbk'))
                if length > max_len: max_len = length
            except:
                pass
        final_width = min(max_len + 4, 60)
        col_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[col_letter].width = final_width
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def merge_and_write(writer, sheet_name, daily_df, old_df, node_order, archive_list):
    """
    合并写入，并执行【自动归档】逻辑
    """
    key_cols = ['环节', '异常类型', '异常描述', '异常信息', '源表异常备注']
    meta_cols = ['优先级', '计划', '责任人', '异常修复', '备注']

    # 1. 数据清洗
    if daily_df is not None and not daily_df.empty:
        for col in key_cols: daily_df[col] = daily_df[col].astype(str)
    if old_df is not None and not old_df.empty:
        if '源表异常备注' not in old_df.columns: old_df['源表异常备注'] = ''
        for col in key_cols: old_df[col] = old_df[col].fillna('').astype(str)

    # 2. 合并
    if old_df is None or old_df.empty:
        if daily_df.empty: return
        final_df = daily_df.copy()
        for col in meta_cols: final_df[col] = ''
    else:
        if daily_df.empty:
            final_df = old_df
        else:
            final_df = pd.merge(old_df, daily_df, on=key_cols, how='outer')
            for col in meta_cols:
                final_df[col] = final_df[col].fillna('') if col in final_df.columns else ''

    # ================= 🚀 自动归档核心逻辑 =================
    if '异常修复' in final_df.columns:
        # 只要包含 "已修复" 或者 "已定位"，都算作解决
        mask_solved = final_df['异常修复'].astype(str).str.contains('已修复|已定位', na=False, regex=True)

        solved_rows = final_df[mask_solved].copy()
        if not solved_rows.empty:
            print(f"   >>> 发现 {len(solved_rows)} 条【已修复/已定位】数据，正在移入归档...")
            archive_list.append(solved_rows)

        final_df = final_df[~mask_solved]
    # ================================================================

    # 3. 排序 (IT异常置顶 -> 环节顺序)
    final_df['环节'] = pd.Categorical(final_df['环节'], categories=node_order, ordered=True)
    final_df['_sort_priority'] = final_df['异常类型'].apply(lambda x: 0 if str(x).strip() == 'IT异常' else 1)
    final_df = final_df.sort_values(by=['_sort_priority', '环节']).drop(columns=['_sort_priority'])

    # 4. 列顺序调整 (固定列 + 日期列)
    fixed_cols = ['环节', '异常类型', '异常描述', '异常信息', '源表异常备注'] + meta_cols
    all_cols = final_df.columns.tolist()

    # 提取所有非固定列（即日期列）
    date_cols = [c for c in all_cols if c not in fixed_cols]

    # 【核心修改】对日期列进行降序排序 (最新的排前面)
    # reverse=True 表示降序 (12/25, 12/24, 12/23...)
    try:
        date_cols.sort(key=lambda x: pd.to_datetime(x, format="%m/%d", errors='coerce'), reverse=True)
    except Exception as e:
        print(f"⚠️ 日期排序出现小问题（不影响数据）: {e}")

    # 重新组合：固定列 + 排序后的日期列
    final_df = final_df[fixed_cols + date_cols]

    final_df.to_excel(writer, sheet_name=sheet_name, index=False)
    adjust_excel_style(writer, sheet_name, final_df)


# ================= 主程序 =================
try:
    current_date = get_current_date_col()
    output_filename = "异常概括汇总.xlsx"
    master_file_path = os.path.join(base_path, output_filename)

    print(f"目标汇总文件: {master_file_path}")

    # 1. 读取新数据
    data_map = {
        '自营采购退货': read_daily_data(files_self_run, current_date),
        '备件库退货': read_daily_data(files_spare_parts, current_date),
        '平台零售订单': read_daily_data(files_platform_retail, current_date)
    }

    # 2. 读取历史 Sheet
    all_old_sheets = load_all_sheets(master_file_path)

    # 准备归档收集器
    all_solved_items = []

    # 3. 写入文件
    with pd.ExcelWriter(master_file_path, engine='openpyxl') as writer:

        # A. 处理业务 Sheet
        for sheet_name, (daily_df, node_order) in data_map.items():
            print(f"正在处理业务 Sheet: [{sheet_name}] ...")
            old_df = all_old_sheets.get(sheet_name)
            merge_and_write(writer, sheet_name, daily_df, old_df, node_order, all_solved_items)

            if sheet_name in all_old_sheets:
                del all_old_sheets[sheet_name]

        # B. 处理“问题归档” Sheet
        archive_df = all_old_sheets.get(archive_sheet_name, pd.DataFrame())

        if all_solved_items:
            new_archive_df = pd.concat(all_solved_items)
            archive_df = pd.concat([archive_df, new_archive_df], ignore_index=True)
            print(f"正在写入归档 Sheet: [{archive_sheet_name}] (新增 {len(new_archive_df)} 条)...")
        else:
            print(f"正在保留归档 Sheet: [{archive_sheet_name}] (无新增)...")

        if not archive_df.empty:
            archive_df.to_excel(writer, sheet_name=archive_sheet_name, index=False)
            adjust_excel_style(writer, archive_sheet_name, archive_df)

            if archive_sheet_name in all_old_sheets:
                del all_old_sheets[archive_sheet_name]

        # C. 处理其他 Sheet
        for sheet_name, other_df in all_old_sheets.items():
            print(f"正在保留其他 Sheet: [{sheet_name}] ...")
            other_df.to_excel(writer, sheet_name=sheet_name, index=False)
            adjust_excel_style(writer, sheet_name, other_df)

    print("=" * 30)
    print(f"✅ 更新完成！")
    print(f"✨ 日期列已调整为【倒序排列】（最新日期紧跟备注列）。")
    print("=" * 30)

except Exception as e:
    import traceback

    print("❌ 发生错误:")
    traceback.print_exc()
    input("按回车键退出...")